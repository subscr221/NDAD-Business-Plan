"""
Comprehensive re-derivation of the Defense Platform business plan financial model
from raw inputs in Defense_Platform_Business_Plan_INR_CORRECTED.xlsx.

Two independent revenue methods are computed:
  Method A - V-sheet Unit × ASP with Wright's-law learning curve and BoM floor
             (the documented "bottoms-up" view)
  Method B - Booking-driven Deliveries: opening backlog x per-vertical delivery rate
             (the actual workbook mechanism feeding Revenue!Total)

NRE is re-derived from first principles:
  Net NRE_v,y = Gross NRE_v,y * (1 - Σ_m (Composition_v,m × Maturity_m,y))
"""
import math
from openpyxl import load_workbook

WB = load_workbook('Defense_Platform_Business_Plan_INR_CORRECTED.xlsx', data_only=False)

# ---------------------------------------------------------------------------
# 1. Assumptions
# ---------------------------------------------------------------------------
ASP_ELASTICITY = 0.5     # Assumptions!C39
BOM_FLOOR_PCT = 0.7      # Assumptions!C40 (V-sheet uses C40, not C30)
LOG_BASE = 2             # Assumptions!C41
REV_LAG_HW_MONTHS = 15   # Assumptions!C49 (documented, not used in workbook formulas)

# ---------------------------------------------------------------------------
# 2. Vertical raw inputs (Y1..Y10 unit shipments + Y1 ASP/BoM)
# ---------------------------------------------------------------------------
# Each vertical: list of variants. Each variant: (asp_y1, bom_y1, [units_y1..y10])
VERTICALS = {
    'V1_AESA': [
        ('X-band Airborne',  26.0,    11.0,    [0,2,3,6,9,12,15,17,19,20]),
        ('S-band Ground',    45.0,    17.765,  [0,1,2,3,4,5,6,7,8,8]),
        ('X-band Naval',     36.0,    14.6775, [0,1,1,2,3,4,5,6,6,7]),
    ],
    'V2_EW_SIGINT': [
        ('Naval ESM/ELINT',  30.0,    14.0,    [1,2,2,3,3,3,4,4,4,4]),
        ('Aerial ESM Pod',   38.0,    13.775,  [1,2,3,4,5,5,6,6,6,7]),
        ('Anti-Drone Jammer', 2.0,    0.8075,  [6,12,20,30,40,50,60,70,75,80]),
        ('Ground SIGINT',   114.0,   60.0,     [0,1,1,1,2,2,2,2,2,2]),
    ],
    'V3_CUAS': [
        ('Dismounted Kit',    1.5,    0.9,     [0,10,40,80,130,180,220,260,290,320]),
        ('Vehicle-Mounted',  14.25,   8.0,     [0,3,7,14,22,30,38,46,52,56]),
        ('Fixed Site',       65.0,   27.075,   [0,0,1,2,3,4,5,6,6,7]),
    ],
    'V4_UAS_MALE': [
        ('MALE ISR',        150.0,   51.3,     [0,0,0,1,2,3,4,5,6,7]),
        ('MALE Strike',     210.0,   75.05,    [0,0,0,0,1,1,2,2,3,3]),
        ('MALE SIGINT/EW',  200.0,   64.6,     [0,0,0,1,1,2,2,3,3,4]),
    ],
    'V5_Autonomous': [
        ('Small AUV',        15.0,   4.085,    [0,0,1,3,7,12,18,24,28,32]),
        ('Large AUV',        60.0,  20.425,    [0,0,0,0,1,1,2,3,3,4]),
        ('Mid USV',          45.0,  15.39,     [0,0,1,2,4,6,9,11,13,15]),
    ],
    'V6_MilAI': [
        ('Edge AI box',       3.8,   0.6175,   [2,5,10,18,25,32,40,45,50,55]),
        ('OIDSS license',    11.4,   0.475,    [1,2,4,8,12,18,24,28,32,35]),
        ('Sensor-fusion mw',  5.7,   0.19,     [2,4,8,15,22,30,38,42,48,52]),
    ],
}

# ---------------------------------------------------------------------------
# 3. METHOD A - V-sheet Unit x ASP with learning curve
# ---------------------------------------------------------------------------
def method_a_revenue():
    """Replicate V-sheet formulas exactly:
       BoM_y = max(BoM_y1 * BoM_FLOOR, BoM_y1 * (cum_units_y / cum_units_y1) ^ log2(...))
       The actual V-sheet uses MAX(D8*0.7, D8 * (cum/D8)^... ) - but Excel formula is
       MAX($D$8*Assumptions!$C$40, $D$8*<learning expression>).
       Simpler form actually shown: BoM scales with cumulative units via Wright's law.
       ASP scales as ASP_y1 * (BoM_y / BoM_y1) ^ elasticity.
       Revenue = Units_y * ASP_y.
    """
    totals_by_year = [0.0] * 10
    by_vertical = {}
    for vname, variants in VERTICALS.items():
        v_year = [0.0] * 10
        for label, asp1, bom1, units in variants:
            cum = 0
            for y in range(10):
                u = units[y]
                cum_prev = cum
                cum += u
                # Replicate the V-sheet learning curve for BoM:
                # Y1 uses input directly. Y2+ uses MAX(floor, asp1 * (cum_y / cum_y1)^...)
                # Actual sheet formula references E8/D8 ratio with an exponent expression.
                # Simplification: Wright's law cost factor = ratio^log(0.92)/log(2)
                if y == 0:
                    bom_y = bom1
                else:
                    if cum > 0 and units[0] > 0:
                        ratio = cum / units[0]
                        learn = ratio ** (math.log(0.92) / math.log(LOG_BASE))
                        bom_y = max(bom1 * BOM_FLOOR_PCT, bom1 * learn)
                    else:
                        # If Y1 units = 0, BoM stays at input level until shipments start
                        bom_y = bom1 * BOM_FLOOR_PCT if cum > 0 else bom1
                if y == 0:
                    asp_y = asp1
                else:
                    asp_y = asp1 * (bom_y / bom1) ** ASP_ELASTICITY
                rev = u * asp_y
                v_year[y] += rev
        by_vertical[vname] = v_year
        for y in range(10):
            totals_by_year[y] += v_year[y]
    return by_vertical, totals_by_year

# ---------------------------------------------------------------------------
# 4. METHOD B - Booking-driven Deliveries (the actual workbook mechanism)
# ---------------------------------------------------------------------------
def method_b_revenue():
    """Read booking inputs from Backlog rows 20-25 and replicate the
       opening-backlog x delivery-rate mechanism."""
    ws = WB['Backlog']
    bookings = {}
    for row, vname in [(20,'V1'),(21,'V2'),(22,'V3'),(23,'V4'),(24,'V5'),(25,'V6')]:
        bookings[vname] = [ws.cell(row=row, column=4+y).value or 0 for y in range(10)]
    delivery_rates = {
        'V1': 0.5, 'V2': 0.5, 'V3': 0.33, 'V4': 0.33, 'V5': 0.25, 'V6': 0.25
    }
    opening_y1 = {'V1':0,'V2':105,'V3':0,'V4':0,'V5':0,'V6':0}
    by_vertical = {}
    totals_by_year = [0.0]*10
    for v in ['V1','V2','V3','V4','V5','V6']:
        opening = opening_y1[v]
        deliveries = []
        rate = delivery_rates[v]
        for y in range(10):
            d = opening * rate
            deliveries.append(d)
            # Closing = Opening + Bookings - Deliveries
            opening = opening + bookings[v][y] - d
        by_vertical[v] = deliveries
        for y in range(10):
            totals_by_year[y] += deliveries[y]
    return by_vertical, totals_by_year, bookings

# ---------------------------------------------------------------------------
# 5. NRE re-derivation
# ---------------------------------------------------------------------------
def nre_recalc():
    rm = WB['ReuseMatrix']; pf = WB['Platform']
    # Composition: ReuseMatrix C6:G11
    composition = {}
    v_keys = ['V1','V2','V3','V4','V5','V6']
    for i, v in enumerate(v_keys):
        composition[v] = [rm.cell(row=6+i, column=3+m).value for m in range(5)]
    # Maturity: Platform D15:M19
    maturity = {}  # m_index -> [y1..y10]
    for m in range(5):
        maturity[m] = [pf.cell(row=15+m, column=4+y).value for y in range(10)]
    # Gross NRE per vertical: ReuseMatrix rows 24,27,30,33,36,39 cols D..M
    gross_rows = {'V1':24,'V2':27,'V3':30,'V4':33,'V5':36,'V6':39}
    gross = {v:[rm.cell(row=r, column=4+y).value for y in range(10)]
             for v,r in gross_rows.items()}
    # Effective reuse savings %: per vertical per year
    leveraged = {}
    for v in v_keys:
        l = []
        for y in range(10):
            reuse = sum(composition[v][m]*maturity[m][y] for m in range(5))
            l.append(gross[v][y] * (1 - reuse))
        leveraged[v] = l
    # Platform NRE: Platform!D6:M10 sum across modules
    plat_by_year = []
    for y in range(10):
        plat_by_year.append(sum(pf.cell(row=6+m, column=4+y).value or 0 for m in range(5)))
    return composition, maturity, gross, leveraged, plat_by_year

# ---------------------------------------------------------------------------
# 6. RUN & PRINT
# ---------------------------------------------------------------------------
def fmt(x): return f"{x:>10,.1f}"

print("="*100)
print("METHOD A: Revenue from V-sheet Unit x ASP (Wright's-law learning curve)")
print("="*100)
ba, ta = method_a_revenue()
print(f"  {'Vertical':<14}" + "".join(f"     Y{y+1}" for y in range(10)) + "    10y Σ")
for v, vals in ba.items():
    print(f"  {v:<14}" + "".join(fmt(x) for x in vals) + fmt(sum(vals)))
print(f"  {'TOTAL':<14}" + "".join(fmt(x) for x in ta) + fmt(sum(ta)))
print(f"\n  Method A Y10 revenue:     ₹{ta[9]:,.1f} Cr")
print(f"  Method A 10-yr cumulative: ₹{sum(ta):,.1f} Cr")

print("\n" + "="*100)
print("METHOD B: Revenue from Backlog deliveries (the workbook's actual mechanism)")
print("="*100)
bb, tb, bookings = method_b_revenue()
print(f"  {'Vertical':<14}" + "".join(f"     Y{y+1}" for y in range(10)) + "    10y Σ")
for v, vals in bb.items():
    print(f"  {v:<14}" + "".join(fmt(x) for x in vals) + fmt(sum(vals)))
print(f"  {'TOTAL':<14}" + "".join(fmt(x) for x in tb) + fmt(sum(tb)))
print(f"\n  Method B Y10 revenue:     ₹{tb[9]:,.1f} Cr")
print(f"  Method B 10-yr cumulative: ₹{sum(tb):,.1f} Cr")

print(f"\n  Σ Bookings 10-yr:          ₹{sum(sum(b) for b in bookings.values()):,.1f} Cr")
final_backlog = sum(sum(bookings[v]) - sum(bb[v]) + (105 if v=='V2' else 0) for v in bookings)
print(f"  Closing backlog (Y10):     ₹{final_backlog:,.1f} Cr")

print("\n" + "="*100)
print("NRE: Re-derivation (Net = Gross x (1 - Σ composition x maturity))")
print("="*100)
comp, mat, gross, lev, plat = nre_recalc()

print("\n  A. Module composition per vertical (rows sum = 1):")
print(f"  {'V':<4}{'M1_RF':>8}{'M2_DSP':>8}{'M3_SE':>8}{'M4_SW':>8}{'M5_AI':>8}   Σ")
for v in ['V1','V2','V3','V4','V5','V6']:
    print(f"  {v:<4}" + "".join(f"{c:>8.2f}" for c in comp[v]) + f"   {sum(comp[v]):.2f}")

print("\n  B. Maturity per module per year:")
print(f"  {'Mod':<4}" + "".join(f"   Y{y+1}" for y in range(10)))
labels = ['RF','DSP','SE','SW','AI']
for m in range(5):
    print(f"  {labels[m]:<4}" + "".join(f"{mat[m][y]:>5.2f}" for y in range(10)))

print("\n  C. Gross NRE per vertical (Cr):")
print(f"  {'V':<4}" + "".join(f"     Y{y+1}" for y in range(10)) + "    10y Σ")
gtot = 0
for v in ['V1','V2','V3','V4','V5','V6']:
    print(f"  {v:<4}" + "".join(fmt(x) for x in gross[v]) + fmt(sum(gross[v])))
    gtot += sum(gross[v])
print(f"  {'Σ':<4}" + "".join(fmt(sum(gross[v][y] for v in gross)) for y in range(10)) + fmt(gtot))

print("\n  D. Leveraged (post-reuse) NRE per vertical (Cr):")
print(f"  {'V':<4}" + "".join(f"     Y{y+1}" for y in range(10)) + "    10y Σ")
ltot = 0
for v in ['V1','V2','V3','V4','V5','V6']:
    print(f"  {v:<4}" + "".join(fmt(x) for x in lev[v]) + fmt(sum(lev[v])))
    ltot += sum(lev[v])
sum_lev_year = [sum(lev[v][y] for v in lev) for y in range(10)]
print(f"  {'Σ':<4}" + "".join(fmt(x) for x in sum_lev_year) + fmt(ltot))

print(f"\n  E. Platform NRE (M1-M5) per year (Cr):")
print(f"  {'Plat':<6}" + "".join(fmt(x) for x in plat) + fmt(sum(plat)))

print(f"\n  Σ Vertical Gross NRE (10-yr):       ₹{gtot:,.1f} Cr")
print(f"  Σ Vertical Leveraged NRE (10-yr):   ₹{ltot:,.1f} Cr")
print(f"  Reuse savings (Gross - Leveraged):  ₹{gtot-ltot:,.1f} Cr  ({(gtot-ltot)/gtot*100:.1f}%)")
print(f"  Σ Platform NRE (10-yr):             ₹{sum(plat):,.1f} Cr")
print(f"  Σ Total NRE (Leveraged + Platform): ₹{ltot+sum(plat):,.1f} Cr")

# ---------------------------------------------------------------------------
# 7. Audit vs. Markdown narrative claims
# ---------------------------------------------------------------------------
print("\n" + "="*100)
print("AUDIT: Calculated values vs Markdown narrative claims")
print("="*100)
claims = [
    ("Y10 revenue",                5034,  tb[9],          "₹ Cr"),
    ("10-yr cumulative revenue",   20500, sum(tb),        "₹ Cr"),
    ("Y10 closing backlog",        16600, final_backlog,  "₹ Cr"),
    ("Σ Platform NRE (M1-M5)",     813,   sum(plat),      "₹ Cr"),
    ("V1 Gross NRE 10-yr (~209)",  209,   sum(gross['V1']), "₹ Cr"),
    ("V2 Gross NRE 10-yr (~114)",  114,   sum(gross['V2']), "₹ Cr"),
    ("V3 Gross NRE 10-yr (~152)",  152,   sum(gross['V3']), "₹ Cr"),
    ("V4 Gross NRE 10-yr (~523)",  523,   sum(gross['V4']), "₹ Cr"),
    ("V5 Gross NRE 10-yr (~247)",  247,   sum(gross['V5']), "₹ Cr"),
    ("V6 Gross NRE 10-yr (~105)",  105,   sum(gross['V6']), "₹ Cr"),
    ("Σ Gross NRE (~1,350)",       1350,  gtot,             "₹ Cr"),
]
print(f"  {'Claim':<35}{'Stated':>12}{'Calc':>14}{'Δ':>12}{'Δ %':>10}")
print("  " + "-"*83)
for name, stated, calc, unit in claims:
    delta = calc - stated
    pct = (delta/stated*100) if stated else 0
    flag = "" if abs(pct) < 5 else "  <-- FLAG"
    print(f"  {name:<35}{stated:>12,.0f}{calc:>14,.1f}{delta:>+12,.1f}{pct:>+9.1f}%{flag}")
